import streamlit as st
import pandas as pd
from lxml import etree
import io
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="REMALAB Dinamik Yönetim", layout="wide", page_icon="⚙️")

# --- YARDIMCI FONKSİYONLAR ---
def normalize_text(text):
    """Metni normalize eder: Türkçe karakterler, boşluk/alt çizgi, büyük-küçük harf."""
    if not text:
        return ""
    text = str(text).strip().lower()
    tr_map = {"ş": "s", "ç": "c", "ğ": "g", "ü": "u", "ö": "o", "ı": "i"}
    for k, v in tr_map.items():
        text = text.replace(k, v)
    # Boşluk, alt çizgi ve tire'yi kaldır — esnek sayfa adı eşleşmesi için
    text = re.sub(r'[\s_\-]+', '', text)
    return text

def clean_date(date_val):
    """Tarih formatını YYYYMMDD şekline zorlar."""
    try:
        if pd.isna(date_val):
            return ""
    except (TypeError, ValueError):
        pass
    clean_val = re.sub(r'[^0-9]', '', str(date_val))
    return clean_val[:8]

# --- 1. GELİŞMİŞ XSD ANALİZ MOTORU ---
def xsd_derin_analiz(xsd_file):
    tree = etree.parse(xsd_file)
    root = tree.getroot()

    # Namespace haritasını düzenle
    ns = {}
    for k, v in root.nsmap.items():
        ns[k if k else 'tns'] = v
    if 'xs' not in ns:
        ns['xs'] = 'http://www.w3.org/2001/XMLSchema'

    # Tüm simpleType enum'larını topla
    all_enums = {}
    for stype in root.xpath('//xs:simpleType', namespaces=ns):
        s_name = stype.get('name')
        vals = [el.get('value') for el in stype.xpath('.//xs:enumeration', namespaces=ns)]
        if s_name and vals:
            all_enums[s_name] = vals

    # Anahtar isimler Excel dosyasındaki gerçek sayfa isimleriyle eşleşmelidir
    # normalize_text ile karşılaştırıldığından Türkçe/boşluk/alt çizgi farkı önemli değil
    yapı = {
        "Genel_Bilgiler": [],
        "Urunler": [],
        "Hammaddeler": []
    }

    skip_names = {"UBFBilgileri", "UBFGenelBilgiler", "UrunBilgileri", "Urun", "HamMaddeBilgileri", "HamMadde"}

    for el in root.xpath('//xs:element', namespaces=ns):
        name = el.get('name')
        if not name or name in skip_names:
            continue

        type_attr = el.get('type', '')
        type_clean = type_attr.split(':')[-1] if ':' in type_attr else type_attr

        # Enum listesini bul
        final_list = all_enums.get(type_clean, [])
        if not final_list:
            inline_vals = [v.get('value') for v in el.xpath('.//xs:enumeration', namespaces=ns)]
            final_list = inline_vals if inline_vals else []
        if not final_list and type_clean:
            final_list = [
                v.get('value')
                for v in root.xpath(f"//xs:simpleType[@name='{type_clean}']//xs:enumeration", namespaces=ns)
            ]

        min_occurs = el.get('minOccurs', '1')
        zorunlu = min_occurs != '0'
        lower_name = name.lower()

        # Üst eleman adını bul
        parent_name = ""
        parent = el.getparent()
        while parent is not None:
            p_name = parent.get('name', '')
            if p_name:
                parent_name = p_name.lower()
                break
            parent = parent.getparent()

        # Hangi sayfaya ait olduğunu belirle
        if any(x in lower_name for x in ["belge", "ruhsat", "tarih", "bolge"]):
            target = "Genel_Bilgiler"
        elif "hammadde" in parent_name or lower_name in ["cins", "mensei", "referansformno", "referansformtipi", "referanssirano"]:
            target = "Hammaddeler"
        else:
            target = "Urunler"

        yapı[target].append({"Ad": name, "Zorunlu": zorunlu, "Liste": final_list})

    return yapı

# --- 2. EXCEL MOTORU ---
def remalab_stili_excel_olustur(yapı, template_path):
    wb = load_workbook(template_path)

    header_fill_red = PatternFill(start_color="FF4B4B", end_color="FF4B4B", fill_type="solid")
    header_fill_grey = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    white_bold = Font(bold=True, color="FFFFFF")
    black_bold = Font(bold=True, color="000000")

    data_sheet_name = "DATA_LISTESI"

    # DATA_LISTESI sayfasını temizle veya oluştur
    if data_sheet_name in wb.sheetnames:
        del wb[data_sheet_name]
    ws_data = wb.create_sheet(title=data_sheet_name)

    # Mevcut tüm named range'leri temizle (L_ ile başlayanlar)
    to_delete = [name for name in wb.defined_names if name.startswith("L_")]
    for name in to_delete:
        del wb.defined_names[name]

    data_col_index = 1

    for sheet_name, columns in yapı.items():
        # Sayfayı normalize ad ile bul
        hedef_sayfa = next(
            (s for s in wb.sheetnames if normalize_text(s) == normalize_text(sheet_name)),
            None
        )
        if not hedef_sayfa:
            continue

        ws = wb[hedef_sayfa]

        # Eski veri satırlarını temizle (başlık satırı korunur)
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row)

        # Eski data validasyonlarını temizle
        ws.data_validations.dataValidation = []

        # Mevcut başlıkları normalize ederek haritala
        mevcut_basliklar = {}
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=1, column=c).value
            if cell_val:
                mevcut_basliklar[normalize_text(cell_val)] = c

        for col_data in columns:
            xsd_ad = col_data["Ad"]
            norm_xsd_ad = normalize_text(xsd_ad)

            # Sütun pozisyonunu bul veya yeni sütun ekle
            target_col = mevcut_basliklar.get(norm_xsd_ad)
            if target_col is None:
                target_col = ws.max_column + 1
                mevcut_basliklar[norm_xsd_ad] = target_col

            cell = ws.cell(row=1, column=target_col)
            if not cell.value:
                cell.value = xsd_ad

            # Başlık stilini ayarla
            if col_data["Zorunlu"]:
                cell.fill = header_fill_red
                cell.font = white_bold
            else:
                cell.fill = header_fill_grey
                cell.font = black_bold

            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_letter = get_column_letter(target_col)

            if col_data["Liste"]:
                # Enum listesini DATA_LISTESI sayfasına yaz
                list_vals = col_data["Liste"]
                for i, val in enumerate(list_vals, start=1):
                    ws_data.cell(row=i, column=data_col_index, value=val)

                # Named range tanımla
                list_name = f"L_{data_col_index}"
                data_col_letter = get_column_letter(data_col_index)
                ref = f"'{data_sheet_name}'!${data_col_letter}$1:${data_col_letter}${len(list_vals)}"
                wb.defined_names[list_name] = DefinedName(list_name, attr_text=ref)

                # Dropdown doğrulaması ekle
                dv_list = DataValidation(
                    type="list",
                    formula1=list_name,
                    allow_blank=True,
                    showDropDown=False
                )
                ws.add_data_validation(dv_list)
                dv_list.add(f"{col_letter}2:{col_letter}1000")
                data_col_index += 1

            elif col_data["Zorunlu"]:
                # Zorunlu alan uyarısı
                dv_warn = DataValidation(
                    type="custom",
                    formula1="=TRUE()",
                    showInputMessage=True
                )
                dv_warn.promptTitle = "REMALAB UYARI"
                dv_warn.prompt = "Bu alan zorunludur!"
                ws.add_data_validation(dv_warn)
                dv_warn.add(f"{col_letter}2:{col_letter}1000")

    # DATA_LISTESI sayfasını gizle
    wb[data_sheet_name].sheet_state = 'hidden'

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

# --- 3. EXCEL → XML DÖNÜŞTÜRÜCÜ ---
def excel_to_xml(excel_file):
    xls = pd.ExcelFile(excel_file)

    def find_sheet(keyword):
        return next((s for s in xls.sheet_names if keyword in normalize_text(s)), None)

    sheet_genel = find_sheet("genel")
    sheet_urun = find_sheet("urun")
    sheet_ham = find_sheet("hammadde")

    eksik = []
    if not sheet_genel: eksik.append("GENEL BİLGİLER")
    if not sheet_urun: eksik.append("URUN_LISTESI")
    if not sheet_ham: eksik.append("HAMMADDE_LISTESI")
    if eksik:
        raise ValueError(f"Şu sayfalar Excel'de bulunamadı: {', '.join(eksik)}")

    # Sayfaları oku, tamamen boş satır/sütunları temizle
    df_genel = xls.parse(sheet_genel).dropna(how='all').reset_index(drop=True)
    df_urun = xls.parse(sheet_urun).dropna(how='all').reset_index(drop=True)
    df_hammadde = xls.parse(sheet_ham).dropna(how='all').reset_index(drop=True)

    # ---- HATA DÜZELTMESİ: Sütun adlarını string'e çevir ----
    df_genel.columns = df_genel.columns.astype(str)
    df_urun.columns = df_urun.columns.astype(str)
    df_hammadde.columns = df_hammadde.columns.astype(str)

    root = etree.Element("UBFBilgileri")
    root.set("xmlns", "http://www.dtm.gov.tr/ubf")

    # 1. Genel Bilgiler
    if df_genel.empty:
        raise ValueError("GENEL BİLGİLER sayfası boş! Lütfen verileri doldurun.")

    genel_node = etree.SubElement(root, "UBFGenelBilgiler")

    # ---- HATA DÜZELTMESİ: .iloc[0] ile ilk satırı al ----
    first_row = df_genel.iloc[0]
    for col in df_genel.columns:
        val = first_row[col]
        if pd.notna(val):
            if "tarih" in col.lower():
                val = re.sub(r'[^0-9]', '', str(val))[:8]
            etree.SubElement(genel_node, str(col)).text = str(val).strip()

    # 2. Ürün Bilgileri
    if df_urun.empty:
        raise ValueError("URUN_LISTESI sayfası boş! Lütfen verileri doldurun.")

    urun_bilgi_node = etree.SubElement(root, "UrunBilgileri")

    for _, urun_row in df_urun.iterrows():
        sira_no_val = str(urun_row.get("SiraNo", "")).strip()
        if not sira_no_val or sira_no_val.lower() == "nan":
            continue

        urun_node = etree.SubElement(urun_bilgi_node, "Urun")

        for col in df_urun.columns:
            if col in ("HamMadde", "SiraNo"):
                continue
            val = urun_row[col]
            if pd.notna(val):
                etree.SubElement(urun_node, str(col)).text = str(val).strip()

        # 3. Hammaddeleri Bağla (SiraNo eşleşmesine göre)
        hm_node = etree.SubElement(urun_node, "HamMaddeBilgileri")

        if not df_hammadde.empty and "SiraNo" in df_hammadde.columns:
            ilgili_hm = df_hammadde[
                df_hammadde["SiraNo"].astype(str).str.strip() == sira_no_val
            ]
            for _, hm_row in ilgili_hm.iterrows():
                hm_item = etree.SubElement(hm_node, "HamMadde")
                for col in df_hammadde.columns:
                    if col == "SiraNo":
                        continue
                    val = hm_row[col]
                    if pd.notna(val):
                        etree.SubElement(hm_item, str(col)).text = str(val).strip()

    return etree.tostring(root, pretty_print=True, encoding="UTF-8", xml_declaration=True)

# --- STREAMLIT ARAYÜZÜ ---
st.sidebar.title("🛠 REMALAB Akıllı Sistemi")
secim = st.sidebar.radio(
    "İşlem seçiniz:",
    ("1. XSD'den Şablon Güncelle", "2. Excel'den XML'e Dönüştür")
)

if secim == "1. XSD'den Şablon Güncelle":
    st.title("📑 REMALAB Tasarımıyla XSD Entegrasyonu")
    st.info("XSD dosyanızı yükleyerek REMALAB.xlsx şablonunu otomatik güncelleyebilirsiniz.")

    uploaded_xsd = st.file_uploader("XSD Dosyasını Seçin", type=["xsd"])

    if uploaded_xsd:
        try:
            with st.spinner("Şablon senkronize ediliyor ve veriler temizleniyor..."):
                yapı = xsd_derin_analiz(uploaded_xsd)
                excel_data = remalab_stili_excel_olustur(yapı, "REMALAB.xlsx")

            st.success("✅ Şablon Güncellendi ve Eski Veriler Temizlendi!")

            # XSD'den bulunan alan sayılarını göster
            col1, col2, col3 = st.columns(3)
            col1.metric("📋 Genel Bilgiler", len(yapı["Genel_Bilgiler"]))
            col2.metric("📦 Ürün Listesi", len(yapı["Urunler"]))
            col3.metric("🧪 Hammadde Listesi", len(yapı["Hammaddeler"]))

            st.download_button(
                label="📥 REMALAB_Guncel.xlsx İndir",
                data=excel_data,
                file_name="REMALAB_Guncel.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except FileNotFoundError:
            st.error("⚠️ 'REMALAB.xlsx' şablon dosyası bulunamadı!")
        except KeyError as e:
            st.error(f"⚠️ Sayfa bulunamadı: {e}. XSD yapısı Excel sayfalarıyla eşleşmiyor.")
        except Exception as e:
            st.error(f"⚠️ Hata: {e}")
            st.exception(e)

elif secim == "2. Excel'den XML'e Dönüştür":
    st.title("🚀 Veri Dönüştürme Merkezi")
    st.info("Doldurduğunuz Excel dosyasını buraya yükleyerek XML çıktısı alabilirsiniz.")

    uploaded_excel = st.file_uploader("Excel Dosyasını Yükleyin", type=["xlsx"])

    if uploaded_excel:
        try:
            with st.spinner("XML Hazırlanıyor..."):
                xml_result = excel_to_xml(uploaded_excel)

            st.success("✅ XML Başarıyla Üretildi!")

            # XML önizleme
            with st.expander("🔍 XML Önizleme (ilk 50 satır)"):
                xml_str = xml_result.decode("utf-8")
                preview_lines = "\n".join(xml_str.splitlines()[:50])
                st.code(preview_lines, language="xml")

            st.download_button(
                label="📥 XML'i İndir",
                data=xml_result,
                file_name="UBF_Veri.xml",
                mime="application/xml"
            )
        except ValueError as e:
            st.error(f"❌ Veri Hatası: {e}")
        except Exception as e:
            st.error(f"❌ Beklenmeyen Hata: {e}")
            st.exception(e)
